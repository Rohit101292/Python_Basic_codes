import openpyxl
from openpyxl import Workbook


wb = openpyxl.load_workbook("/Users/rohitdadala/PycharmProjects/databadeex/venv/file_example_XLSX_5000.xlsx")
sheet = wb.active
rows = sheet.max_row
columns = sheet.max_column

list =[]
for i in sheet.iter_rows(1, rows+1):
    list.append(i[0].value)

data= {}
for i, row in enumerate(sheet.iter_rows(values_only=True)):
    if i == 0:
        data[row[1]] = []
        data[row[2]] = []
        data[row[3]] = []
        data[row[4]] = []
        data[row[5]] = []
        data[row[6]] = []
        data[row[7]] = []

    else:
        data['First Name'].append(row[1])
        data['Last Name'].append(row[2])
        data['Gender'].append(row[3])
        data['Country'].append(row[4])
        data['Age'].append(row[5])
        data['Date'].append(row[6])
        data['Id'].append(row[7])

print(data)

